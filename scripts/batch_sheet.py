"""The progress-spreadsheet machinery shared by the batch scripts.

Extracted from wjazz_batch.py when benchmark_batch.py needed the identical
behaviour for a second sheet — one copy, parameterised by path, headers and
key column, because the failure modes here were all found the hard way:

- Rows are rewritten by header NAME, never by position, so widening a sheet
  never costs the rows already scored (migrate_columns).
- An Excel Table definition on the sheet must be rebuilt to match the grid on
  every save, or Excel reports "We found a problem with some content"
  (sync_table).
- Excel holds an exclusive lock while a sheet is open; both the load and the
  save say "close Excel" by name instead of letting a zipfile traceback out
  of openpyxl (save_sheet / load_or_create_sheet).
"""

from pathlib import Path


def migrate_columns(ws, headers: list[str]) -> bool:
    """Bring a sheet written by an older version up to the current columns.

    Rows are rewritten by header NAME, never by position, so a sheet from
    before a column existed keeps every number already in it and simply gains
    blank cells. Returns whether anything moved, so the caller can say so.
    """
    header = [cell.value for cell in ws[1]] if ws.max_row else []
    if header == headers:
        return False
    records = [
        dict(zip(header, values, strict=False))
        for values in ws.iter_rows(min_row=2, values_only=True)
    ]
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for record in records:
        ws.append([record.get(name) if record.get(name) is not None else "" for name in headers])
    return True


def sync_table(ws, headers: list[str]) -> None:
    """Keep an Excel Table on the sheet matching the grid underneath it.

    A sheet someone has formatted as a Table in Excel carries a definition
    listing its columns BY NAME and the exact range it covers, and openpyxl
    preserves that definition verbatim. Widening the sheet without updating
    it leaves Excel reading a table whose columns and range disagree with the
    data — which is what "We found a problem with some content" means. The
    name and style the user picked are kept; only columns and range rebuild.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import AutoFilter
    from openpyxl.worksheet.table import Table, TableColumn

    if not ws.tables:
        return
    ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    # Keys, not .items(): openpyxl's TableList.items() yields (name, ref
    # STRING) rather than the Table itself.
    for name in list(ws.tables):
        style = ws.tables[name].tableStyleInfo
        del ws.tables[name]
        rebuilt = Table(
            displayName=name,
            name=name,
            ref=ref,
            autoFilter=AutoFilter(ref=ref),
            tableColumns=[
                TableColumn(id=index, name=header) for index, header in enumerate(headers, start=1)
            ],
        )
        rebuilt.tableStyleInfo = style
        ws.add_table(rebuilt)


def save_sheet(wb, ws, sheet_path: Path, headers: list[str]) -> None:
    """The only way a batch writes its sheet — table kept in step.

    A PermissionError here means the spreadsheet is open in Excel, which
    takes an exclusive lock. Worth catching by name: the cause is not in the
    stack, and the run may have spent minutes of CREPE reaching this line.
    """
    sync_table(ws, headers)
    try:
        wb.save(sheet_path)
    except PermissionError as exc:
        raise SystemExit(
            f"cannot write {sheet_path.name}: it is open in another program "
            f"(Excel takes an exclusive lock, and leaves a ~${sheet_path.name} "
            f"beside it while it holds one).\nClose it and re-run.\n  {sheet_path}"
        ) from exc


def load_or_create_sheet(
    sheet_path: Path, headers: list[str], sheet_title: str, seed_rows=None, log=print
):
    """The workbook and worksheet, migrated to the current columns.

    `seed_rows` (an iterable of per-row dicts keyed by header name) fills a
    brand-new sheet so every expected row exists from the first run — the
    wjazzd sheet seeds all 456 melids this way.
    """
    from openpyxl import Workbook, load_workbook

    if sheet_path.is_file():
        try:
            wb = load_workbook(sheet_path)
        except PermissionError as exc:
            raise SystemExit(
                f"cannot read {sheet_path.name}: it is open in another program "
                f"(Excel takes an exclusive lock).\nClose it and re-run.\n  {sheet_path}"
            ) from exc
        ws = wb.active
        if migrate_columns(ws, headers):
            log(f"  (migrated {sheet_path.name} to {len(headers)} columns; rows kept)")
        # Saved unconditionally: an existing sheet may carry a stale table
        # definition from before sync_table existed.
        save_sheet(wb, ws, sheet_path, headers)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in seed_rows or []:
        ws.append([row.get(name) if row.get(name) is not None else "" for name in headers])
    sheet_path.parent.mkdir(parents=True, exist_ok=True)
    save_sheet(wb, ws, sheet_path, headers)
    return wb, ws


def row_index(ws) -> dict:
    """key-cell value -> 1-based worksheet row, from column A, header skipped."""
    index = {}
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None and value != "":
            index[value] = row
    return index


def write_row(ws, index: dict, key, row: dict, fields: list[str]) -> None:
    """Update exactly this key's row, leaving every other row untouched.

    `fields` name the row dict's keys in column order (they may differ from
    the sheet headers — the wjazzd sheet says `number` where the row says
    `melid`). The key must be the row's own column-A value.
    """
    if key is None or key == "":
        return  # nothing identifiable — the caller already logged why
    target = index.get(key)
    if target is None:
        target = ws.max_row + 1
        index[key] = target
    for col, field in enumerate(fields, start=1):
        ws.cell(row=target, column=col, value=row.get(field, ""))
